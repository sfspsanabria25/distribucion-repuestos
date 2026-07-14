import streamlit as st
import openpyxl
import plotly.express as px
from io import BytesIO
from datetime import datetime
from collections import Counter

# app soporte
st.set_page_config(page_title="Distribución de Repuestos", layout="wide")
st.title("📦 Distribución de Repuestos soporte técnico")

# ---------- BOTÓN REINICIAR ----------
if st.button("🔄 Reiniciar"):
    st.session_state.clear()
    st.rerun()

archivo = st.file_uploader("Cargar archivo Excel", type=["xlsx"])


def buscar_columna(encabezados, posibles):
    for p in posibles:
        for e in encabezados:
            if e and p.lower() in str(e).lower():
                return encabezados.index(e)
    return None


# ---------- CONVERTIR FECHA ----------
def convertir_fecha(valor):

    if isinstance(valor, datetime):
        return valor

    try:
        texto = str(valor).split(",")[0].strip()
        return datetime.strptime(texto, "%d/%m/%Y")

    except:
        return datetime.max


if archivo:

    wb = openpyxl.load_workbook(archivo)
    ws = wb.active

    encabezados = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]

    col_caso = buscar_columna(encabezados, ["caso"])
    col_centro = buscar_columna(encabezados, ["centro"])
    col_fecha = 10  # Columna K

    if None in (col_caso, col_centro):
        st.error("❌ Archivo no válido")
        st.stop()

    datos_originales = []
    casos_vistos = set()
    datos = []

    casos_woden = set()
    casos_logy = set()

    # ---------- ELIMINAR DUPLICADOS ----------
    for fila in ws.iter_rows(min_row=2, values_only=True):

        caso = fila[col_caso]
        centro = str(fila[col_centro]).upper() if fila[col_centro] else ""

        # 🔥 TODAS las solicitudes para gráfica
        datos_originales.append(list(fila))

        # 🔥 Duplicados eliminados SOLO para distribución
        if not caso or caso in casos_vistos:
            continue

        casos_vistos.add(caso)

        if "WODEN" in centro:
            casos_woden.add(caso)

        if "LOGYTECH" in centro:
            casos_logy.add(caso)

        datos.append(list(fila))

    total_rep_original = len(datos_originales)
    total_casos = len(datos)

    # ---------- MÉTRICAS ----------
    c1, c2, c3, c4 = st.columns(4)

    c1.metric("Casos únicos", total_casos)
    c2.metric("Casos WODEN", len(casos_woden))
    c3.metric("Casos LOGYTECH", len(casos_logy))
    c4.metric("Solicitudes totales", total_rep_original)

    # ---------- INPUTS ----------
    colA, colB = st.columns(2)

    personas = colA.number_input(
        "Número de líderes técnicos",
        min_value=1,
        step=1
    )

    por_persona = colB.number_input(
        "Casos por líder técnico",
        min_value=1,
        step=1
    )

    total_asignar = personas * por_persona

    st.write(f"📦 Total de casos a asignar: {total_asignar}")

    # ---------- SELECTOR ----------
    modo = st.selectbox(
        "Modo de organización",
        [
            "Modo 1: Prioridad → Casos antiguos - Fecha de solicitud",
            "Modo 2: Prioridad → Fecha de solicitud - Casos antiguos",
            "Modo 3: Prioridad → Fecha de solicitud"
        ]
    )

    # ---------- GRÁFICA INTERACTIVA ----------
    st.subheader("📈 Solicitudes de repuestos por fecha")

    fechas_validas = []

    # 🔥 USAR TODAS LAS SOLICITUDES
    for fila in datos_originales:

        fecha = convertir_fecha(fila[col_fecha])

        if fecha != datetime.max:
            fechas_validas.append(fecha.strftime("%d/%m/%Y"))

    conteo_fechas = Counter(fechas_validas)

    fechas_ordenadas = sorted(
        conteo_fechas.keys(),
        key=lambda x: datetime.strptime(x, "%d/%m/%Y")
    )

    cantidades = [conteo_fechas[f] for f in fechas_ordenadas]

    fig = px.line(
        x=fechas_ordenadas,
        y=cantidades,
        markers=True,
        labels={
            "x": "Fecha de solicitud",
            "y": "Cantidad de solicitudes"
        }
    )

    fig.update_traces(mode="lines+markers")

    fig.update_layout(
        xaxis_title="Fecha de solicitud",
        yaxis_title="Cantidad de solicitudes",
        hovermode="x unified"
    )

    st.plotly_chart(fig, use_container_width=True)

    # ---------- VALIDACIÓN ----------
    if total_casos >= total_asignar:

        st.success(
            f"Datos suficientes. "
            f"Sobrantes estimados: {total_casos - total_asignar}"
        )

    else:
        st.warning("Se asignarán todos los casos disponibles.")

    # ---------- GENERAR ----------
    if st.button("Generar distribución"):

        # ---------- SELECCIÓN SEGÚN MODO ----------
        if "Modo 1" in modo:

            datos.sort(
                key=lambda x: int(x[col_caso])
            )

        else:

            datos.sort(
                key=lambda x: convertir_fecha(x[col_fecha])
            )

        # ---------- SEPARAR PRIORIDAD ----------
        prioridad = []
        normales = []

        for fila in datos:

            centro = str(fila[col_centro]).upper() if fila[col_centro] else ""

            if "WODEN" in centro or "LOGYTECH" in centro:
                prioridad.append(fila)

            else:
                normales.append(fila)

        # ---------- DISTRIBUCIÓN ----------
        grupos = [[] for _ in range(personas)]

        # Prioridad primero
        for i, fila in enumerate(prioridad):
            grupos[i % personas].append(fila)

        # Completar con normales
        indice = 0

        for fila in normales:

            intentos = 0

            while (
                len(grupos[indice % personas]) >= por_persona
                and intentos < personas
            ):
                indice += 1
                intentos += 1

            if intentos >= personas:
                break

            grupos[indice % personas].append(fila)
            indice += 1

        # ---------- SOBRANTES ----------
        asignados_reales = [fila for grupo in grupos for fila in grupo]

        sobrantes = [
            f for f in datos
            if f not in asignados_reales
        ]

        # ---------- ARCHIVO PRINCIPAL ----------
        wb_out = openpyxl.Workbook()
        wb_out.remove(wb_out.active)

        for i, grupo in enumerate(grupos):

            prioridad_local = []
            resto = []

            for fila in grupo:

                centro = str(fila[col_centro]).upper() if fila[col_centro] else ""

                if "WODEN" in centro or "LOGYTECH" in centro:
                    prioridad_local.append(fila)

                else:
                    resto.append(fila)

            # ---------- ORGANIZACIÓN ----------
            if "Modo 1" in modo:

                resto_ordenado = sorted(
                    resto,
                    key=lambda x: int(x[col_caso])
                )

                mitad = len(resto_ordenado) // 2

                primera = resto_ordenado[:mitad]

                segunda = sorted(
                    resto_ordenado[mitad:],
                    key=lambda x: convertir_fecha(x[col_fecha])
                )

                organizados = prioridad_local + primera + segunda

            elif "Modo 2" in modo:

                resto_fecha = sorted(
                    resto,
                    key=lambda x: convertir_fecha(x[col_fecha])
                )

                mitad = len(resto_fecha) // 2

                primera = resto_fecha[:mitad]

                segunda = sorted(
                    resto_fecha[mitad:],
                    key=lambda x: int(x[col_caso])
                )

                organizados = prioridad_local + primera + segunda

            elif "Modo 3" in modo:

                resto_fecha = sorted(
                    resto,
                    key=lambda x: convertir_fecha(x[col_fecha])
                )

                organizados = prioridad_local + resto_fecha

            # ---------- CREAR HOJA ----------
            ws_out = wb_out.create_sheet(f"Tec_lid{i+1}")

            ws_out.append(encabezados)

            for fila in organizados:
                ws_out.append(fila)

        buffer1 = BytesIO()
        wb_out.save(buffer1)

        # ---------- ARCHIVO SOBRANTES ----------
        if "Modo 1" in modo:

            sobrantes.sort(
                key=lambda x: int(x[col_caso]),
                reverse=True
            )

        else:

            sobrantes.sort(
                key=lambda x: convertir_fecha(x[col_fecha]),
                reverse=True
            )

        wb_rest = openpyxl.Workbook()
        ws_rest = wb_rest.active

        ws_rest.title = "Repuestos no asignados"

        ws_rest.append(encabezados)

        for fila in sobrantes:
            ws_rest.append(fila)

        buffer2 = BytesIO()
        wb_rest.save(buffer2)

        # ---------- GUARDAR EN MEMORIA ----------
        st.session_state["dist"] = buffer1.getvalue()
        st.session_state["sobrantes"] = buffer2.getvalue()

    # ---------- DESCARGAS ----------
    if "dist" in st.session_state:

        st.success("Distribución generada ✅")

        st.download_button(
            "⬇️ Descargar distribución",
            data=st.session_state["dist"],
            file_name="Distribucion_Casos.xlsx"
        )

        st.download_button(
            "⬇️ Descargar repuestos no asignados",
            data=st.session_state["sobrantes"],
            file_name="Repuestos_No_Asignados.xlsx"
        )
